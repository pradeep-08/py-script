import os
import openpyxl

MAPPING_RULES = [
    (["voltage to 12v", "battery voltage", "power up"], ["Set_Bat_Vol(12);", "Read_MEC();"]),
    (["mec equal to zero", "read mec"], ["Read_MEC();"]),
    (["extended diagnostic session", "extended session"], ["Extended_Session_Phy();"]),
    (["tester present"], ["Tester_Present_Per_Start_Fun(2000);"]),
    (["disable communication"], ["Dis_Comm_phy();"]),
    (["security level 03", "security level 3"], ["Security_Unlock_With_Key_lvl_3_Phy();"]),
    (["security level 01", "security level 1"], ["Security_Unlock_With_Key_lvl_1_Phy();"]),
    (["security level 05", "security level 5"], ["Security_Unlock_With_Key_lvl_5_Phy();"]),
    (["security level 09", "security level 9"], ["Security_Unlock_With_Key_lvl_9_Phy();"]),
    (["security level 11"], ["Security_Unlock_With_Key_lvl_11_Phy();"]),
    (["security level 13"], ["Security_Unlock_With_Key_lvl_13_Phy();"]),
    (["security level 15"], ["Security_Unlock_With_Key_lvl_15_Phy();"]),
    (["service id 31 and sub function 01", "rid 21e"], ["Write_RID_21E();"]),
    (["ecu reset", "hard reset"], ["Hard_Reset_Phy();"]),
    (["default session"], ["Default_Session_Phy();"]),
    (["normal mode", "back to normal", "save the evidence", "ecu locked"], ["Back_To_Normal();"]),
    (["power mode to off"], ["@IgnitionSwitch=0;", "Read_PowerMode();"]),
    (["power mode to acc"], ["@IgnitionSwitch=1;", "Read_PowerMode();"]),
    (["power mode to run"], ["@IgnitionSwitch=2;", "Read_PowerMode();"]),
    (["power mode to start"], ["@IgnitionSwitch=3;", "Read_PowerMode();"]),
    (["ignition cycle"], ["Set_Bat_Vol(0);", "testWaitForTimeout(1000);", "Set_Bat_Vol(12);"]),
    (["termination step", "stop logging", "saved the evidence", "sba ticket"], ["stopLogging(\"Test Logs\");"])
]

import re

def get_function_body_from_master(master_dir, function_name):
    base_name = function_name.replace("();", "").strip()
    search_pattern = re.compile(r'void\s+' + re.escape(base_name) + r'\s*\([^)]*\)\s*\{', re.MULTILINE)
    
    if not os.path.isdir(master_dir):
        return None
        
    for filename in os.listdir(master_dir):
        if not filename.endswith(('.cin', '.can')):
            continue
            
        filepath = os.path.join(master_dir, filename)
        try:
            with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
                content = f.read()
                match = search_pattern.search(content)
                if match:
                    start_idx = match.end()
                    brace_count = 1
                    end_idx = start_idx
                    while brace_count > 0 and end_idx < len(content):
                        if content[end_idx] == '{':
                            brace_count += 1
                        elif content[end_idx] == '}':
                            brace_count -= 1
                        end_idx += 1
                    return content[start_idx:end_idx-1]
        except Exception:
            pass
    return None

def parse_can_string(can_str, pad_tx=False, is_tx=False):
    if not can_str:
        return "0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00"
    tokens = str(can_str).strip().split()
    
    if is_tx and len(tokens) > 5:
        data = tokens[5:13]
    elif len(tokens) >= 8:
        data = tokens[-8:]
    else:
        data = tokens
        
    while len(data) < 8:
        data.append("55" if pad_tx else "00")
            
    formatted = []
    for d in data:
        if pad_tx and d == "00":
            formatted.append("0x55")  
        else:
            formatted.append(f"0x{d}")
    return ",".join(formatted)

def is_hex_line(line):
    """Check if a line consists of space-separated hex byte tokens (e.g. '50 94 DA 59 F1 03').
    Also allows XX as a placeholder. Skips lines that contain regular text/words."""
    line = line.strip()
    if not line:
        return False
    tokens = line.split()
    # Need at least 3 tokens to look like a CAN message line
    if len(tokens) < 3:
        return False
    for t in tokens:
        # Allow hex bytes (2 chars) and XX placeholder, also != prefixed like !=FF
        if re.match(r'^[0-9a-fA-F]{2}$', t) or t.upper() == 'XX':
            continue
        else:
            return False
    return True

def extract_hex_lines_from_text(text):
    """Extract lines that look like hex CAN data from a mixed text/hex string."""
    if not text:
        return []
    hex_lines = []
    for line in str(text).strip().splitlines():
        line = line.strip()
        if is_hex_line(line):
            hex_lines.append(line)
    return hex_lines

def split_into_frames(tokens):
    """Split a token list that may contain multiple concatenated CAN frames.
    Detects the repeating 5-byte header pattern (e.g. 50 94 DA 59 F1) and splits at each occurrence.
    Returns a list of token lists, one per frame."""
    if len(tokens) <= 13:  # Single frame at most (5 header + 8 data)
        return [tokens]
    
    # Use first 5 tokens as the header pattern
    header = [t.upper() for t in tokens[:5]]
    frames = []
    current_start = 0
    
    i = 5
    while i <= len(tokens) - 5:
        # Check if tokens[i:i+5] matches the header
        if [t.upper() for t in tokens[i:i+5]] == header:
            frames.append(tokens[current_start:i])
            current_start = i
            i += 5
        else:
            i += 1
    
    frames.append(tokens[current_start:])
    return frames

def parse_multiline_can(can_str, pad_tx=False):
    """Parse multi-line CAN messages (e.g. ISO-TP multi-frame with 10/21/22 sequences).
    Each line like '50 94 DA 59 F1 10 0E 27 02 00 00 00 00' is split at the 6th byte
    (after the 5-byte header) to produce individual send_on_CAN / Expect byte strings.
    Handles both actual newlines AND concatenated frames on a single line.
    Skips any lines that are not valid hex byte sequences (e.g. text descriptions).
    Returns a list of formatted byte strings."""
    if not can_str:
        return ["0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00"]
    
    lines = str(can_str).strip().splitlines()
    results = []
    
    for line in lines:
        line = line.strip()
        if not line:
            continue
        
        # Skip lines that don't look like hex data (e.g. "Positive response should receive :")
        if not is_hex_line(line):
            continue
        
        tokens = line.split()
        
        # Split into individual frames if multiple are concatenated on one line
        frames = split_into_frames(tokens)
        
        for frame_tokens in frames:
            # If frame has more than 5 tokens, treat first 5 as header and take from 6th byte onward
            if len(frame_tokens) > 5:
                data = frame_tokens[5:13]
            elif len(frame_tokens) >= 8:
                data = frame_tokens[-8:]
            else:
                data = frame_tokens
            
            # Pad to 8 bytes
            while len(data) < 8:
                data.append("55" if pad_tx else "00")
            
            formatted = []
            for d in data:
                if pad_tx and d == "00":
                    formatted.append("0x55")
                else:
                    formatted.append(f"0x{d}")
            results.append(",".join(formatted))
    
    return results if results else ["0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00"]

def map_step_to_capl(description, tx_val=None, rx_val=None, master_dir=None):
    desc_lower = description.lower()
    mapped_lines = []
    
    tx_bytes = parse_can_string(tx_val, pad_tx=True, is_tx=True)
    rx_bytes = parse_can_string(rx_val, pad_tx=False, is_tx=False)

    for keywords, capl_codes in MAPPING_RULES:
        if any(kw in desc_lower for kw in keywords):
            for code in capl_codes:
                if code not in mapped_lines:
                    if code.endswith("();") and master_dir and code != "Back_To_Normal();":
                        body = get_function_body_from_master(master_dir, code)
                        if body:
                            first_line_desc = description.split('\n')[0].strip()
                            # Replace ACTION
                            body = re.sub(r'ACTION\([^)]*\);\s*(//.*)?', f'ACTION("{first_line_desc}"); // {first_line_desc}', body)
                            
                            # Replace send_on_CAN if we have tx_val
                            if tx_val: 
                                body = re.sub(r'(send_on_CAN(?:_FUN)?)\([^)]*\)', r'\1(' + tx_bytes + ')', body)
                                
                            # Replace Expect (but not Expect1 etc) if rx_val
                            if rx_val:
                                body = re.sub(r'(Expect(?:[A-Z0-9_]*)?)\([^)]*\)', r'\1(' + rx_bytes + ')', body)
                            
                            # Update EXPECTED_DATA and OBSERVED_DATA to use 0X8 instead of 8
                            body = re.sub(r'(EXPECTED_DATA\([^,]+,\s*E_Resp\s*,\s*)8(\s*\))', r'\g<1>0X8\2', body)
                            body = re.sub(r'(OBSERVED_DATA\([^,]+,\s*O_Resp\s*,\s*)8(\s*\))', r'\g<1>0X8\2', body)

                            mapped_lines.extend([line.strip('\r\n') for line in body.split('\n') if line.strip('\r\n')])
                            continue
                            
                    mapped_lines.append(code)
    return mapped_lines

def extract_steps_from_workbook(excel_path, master_dir):
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    sheet = wb.active
    
    desc_col = 2
    step_col = 1
    id_col = None
    tx_col = None
    rx_col = None
    
    # Try to dynamically find columns
    for row in sheet.iter_rows(min_row=1, max_row=5):
        for cell in row:
            if cell.value:
                val = str(cell.value).lower().strip()
                if "description" in val:
                    desc_col = cell.column
                elif val == "step":
                    step_col = cell.column
                elif val == "id":
                    id_col = cell.column
                elif "binary" in val and "expect" not in val and "except" not in val:
                    tx_col = cell.column
                elif "expect" in val or "except" in val:
                    rx_col = cell.column

    steps = []
    unmatched_rows = []
    
    total = 0
    mapped = 0
    first_id = None

    for row_idx in range(2, sheet.max_row + 1):
        desc = sheet.cell(row=row_idx, column=desc_col).value
        step_name = sheet.cell(row=row_idx, column=step_col).value if step_col else None
        
        step_id = sheet.cell(row=row_idx, column=id_col).value if id_col else None
        tx_val = sheet.cell(row=row_idx, column=tx_col).value if tx_col else None
        rx_val = sheet.cell(row=row_idx, column=rx_col).value if rx_col else None
        
        if step_id and not first_id:
            first_id = str(step_id).strip()

        if not step_name:
            step_name = f"Step {row_idx - 1}"
            
        if desc:
            total += 1
            capl_commands = map_step_to_capl(str(desc), tx_val, rx_val, master_dir)
            
            if capl_commands:
                mapped += 1
                steps.append((step_name, str(desc), capl_commands))
            else:
                unmatched_rows.append(f"{step_name}: {str(desc)[:50]}...")
                
                clean_desc = str(desc).split('\n')[0].strip().replace('"', '\\"')
                use_tx = tx_val
                
                # If TX column is empty, try to extract hex lines from the description
                if not use_tx:
                    desc_hex_lines = extract_hex_lines_from_text(str(desc))
                    if desc_hex_lines:
                        use_tx = '\n'.join(desc_hex_lines)
                    
                # Use multiline parser for multi-frame CAN messages
                tx_lines = parse_multiline_can(use_tx, pad_tx=True)
                rx_lines = parse_multiline_can(rx_val, pad_tx=False)
                
                default_capl = [
                    "STEP();",
                    f'ACTION("{clean_desc}"); // {clean_desc}',
                ]
                
                # Add send_on_CAN for each TX frame
                for tx_bytes in tx_lines:
                    default_capl.append(f"send_on_CAN({tx_bytes});")
                    default_capl.append("testWaitForTimeout(50);")
                
                # Add Expect for each RX frame
                for rx_bytes in rx_lines:
                    default_capl.append(f"Expect({rx_bytes});")
                
                default_capl.extend([
                    'EXPECTED_DATA("",E_Resp,0X8);',
                    'OBSERVED_DATA("",O_Resp,0X8);',
                    "Check(E_Resp,O_Resp);  ",
                    "Clear_Buffer();   ",
                    "",
                    "testWaitForTimeout(1000);"
                ])
                steps.append((step_name, str(desc), default_capl))

    # If it's totally empty or we couldn't parse, let's at least not break
    unmatched = total - mapped
    
    return steps, total, mapped, unmatched, unmatched_rows, first_id

def build_testcase_from_steps(testcase_name, steps):
    code = f"testcase {testcase_name}()\n{{\n"
    code += f'  char logPath[256] = "C:\\\\Users\\\\Logs\\\\{testcase_name}.asc";\n'
    code += '  setLogFileName("Test Logs",logPath);\n'
    code += '  startLogging("Test Logs");\n\n'
    
    for i, (step_name, desc, capl_commands) in enumerate(steps, 1):
        code += f"  /*-------------------------------Test Step {i} ---------------------------------------*/\n"
        clean_desc = desc.replace('\n', ' ')
        code += f"  /* ACTION: {clean_desc} */\n"
        
        # Add a newline if it's a termination step to match user's requested formatting
        if any("stopLogging" in cmd for cmd in capl_commands):
            code += "\n"
            
        for capl in capl_commands:
            code += f"  {capl}\n"
        code += "\n"
        
    # Only add stopLogging at the end if it wasn't already included in the last step
    last_step_has_stop = False
    if steps:
        last_step_cmds = steps[-1][2]
        if any("stopLogging" in c for c in last_step_cmds):
            last_step_has_stop = True
            
    if not last_step_has_stop:
        code += "  stopLogging(\"Test Logs\");\n"
    code += "}\n"
    return code

def generate_can_from_excel_with_master(excel_path, master_can_path, output_dir):
    try:
        master_dir = os.path.dirname(master_can_path)
        # 1. Parse Excel
        steps, total, mapped, unmatched, unmatched_arr, first_id = extract_steps_from_workbook(excel_path, master_dir)
        
        # 2. Build the exact testcase code
        base_name = os.path.basename(excel_path).replace('.xlsx', '').replace('.xls', '')
        
        # Determine testcase name based on first ID found, falling back to file base_name
        tc_seed = first_id if first_id else base_name
        safe_tc_name = "TC_Generated_" + "".join(c if c.isalnum() else "_" for c in tc_seed)
        
        testcase_code = build_testcase_from_steps(safe_tc_name, steps)
        
        # 3. Combine with master (we extract the exact includes/variables from your master)
        final_code = "/* Generated CAPL code string here */\n"
        final_code += f"/* Source: {os.path.basename(excel_path)} */\n"
        
        if os.path.exists(master_can_path):
            # Just read the first lines (includes, variables) from the master snippet
            with open(master_can_path, "r", encoding="utf-8") as m_file:
                content = m_file.read()
                # find the first testcase and extract everything before it
                idx = content.find("testcase ")
                if idx != -1:
                    final_code += content[:idx]
        
        # Append our completely customized testcase
        final_code += "\n\n" + testcase_code
        
        # Output saving logic if we wanted to save it locally
        output_path = os.path.join(output_dir, f"{safe_tc_name}.can")
        with open(output_path, "w", encoding="utf-8") as out:
            out.write(final_code)

        return {
            "total": total,
            "mapped": mapped,
            "unmatched": unmatched,
            "unmatchedSteps": unmatched_arr,
            "previewCode": final_code
        }
    except Exception as e:
        print(f"Error during generation: {e}")
        return {
            "total": 0,
            "mapped": 0,
            "unmatched": 0,
            "unmatchedSteps": [str(e)],
            "previewCode": f"// Generation failed: {str(e)}"
        }
